from openpyxl import load_workbook
from openpyxl import Workbook

def get_rules(filename):
    file = open(filename, 'r')
    key_column = file.readline().split(':')[1]
    key = file.readline().split(':')[1]

    targets = []
    for tmp_line in file:
        line = tmp_line.strip()
        wb_name = line.split(':')[0]
        ws_names = line.split(':')[1].split(',')

        target = {wb_name : ws_names}
        targets.append(target)

    return key_column.strip(), key.strip(), targets

def find_column(ws, key_column):
    column = -1
    for col in range(1, ws.max_column+1):
        if ws.cell(row=1, column=col).value == key_column:
            # print(ws.cell(row=1, column=col).value)
            column = col
            break
    return column

def get_rows(ws, key, column):
    rows = []

    for row in range(1, ws.max_row+1):
        if ws.cell(row=row, column=column).value == int(key):
            rows.append(ws[row])
    
    return rows

def get_raw_row(row):
    cells = []

    for cell in row:
        cells.append(cell.value)
    
    return cells

def get_row_by(key_column, key, target):
    wb_name = list(target.keys())[0]
    ws_names = target[wb_name]

    rows = []
    wb = load_workbook(wb_name)
    for ws_name in ws_names:
        ws = wb[ws_name]
        column = find_column(ws, key_column)

        if column != -1:
            rows_cell = get_rows(ws, key, column)
            for row in rows_cell:
                rows_raw = get_raw_row(row)
                rows.append(rows_raw)
    return rows

def main():
    key_column, key, targets = get_rules('규칙파일.txt')

    rows = []

    merge_info_wb = Workbook()
    active_ws = merge_info_wb.active
    active_ws_title = '취합_' + key

    for target in targets:
        tmp_rows = get_row_by(key_column, key, target)
        
        for row in tmp_rows:
            active_ws.append(row)
    
    merge_info_wb.save('취합_정보_' + key + '.xlsx')

if __name__ == "__main__":
    main()
